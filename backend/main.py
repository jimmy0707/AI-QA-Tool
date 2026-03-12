from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import pandas as pd
import requests
import json
import os
import io
import uuid
import logging
import psutil
import platform
import asyncio
import time
import threading
import collections
from typing import Optional, List
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import OpenAI, RateLimitError, AuthenticationError, APIError
from google import genai
from google.genai import types as genai_types
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from ai_rate_manager import ai_manager

# Windows-safe logging — do NOT use basicConfig as it conflicts with uvicorn on Windows
logger = logging.getLogger("ai-qa-platform")
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(_handler)
    logger.setLevel(logging.INFO)
    logger.propagate = False

app = FastAPI(title="AI QA Decision Intelligence Platform", version="1.0")

# Initialize AI rate manager on startup
@app.on_event("startup")
async def startup():
    await ai_manager.initialize()
    logger.info("AI Rate Manager ready")

# IMPROVEMENT 5: CORS origins from environment (secure)
_raw_origins = os.getenv("ALLOWED_ORIGINS", "*")
ALLOWED_ORIGINS = ["*"] if _raw_origins == "*" else [o.strip() for o in _raw_origins.split(",")]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
OLLAMA_MODEL_DEFAULT = os.getenv("OLLAMA_MODEL", "phi3")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# Ollama Model Registry
# ─────────────────────────────────────────────
OLLAMA_MODELS = {
    "tinyllama": {
        "label": "TinyLlama",
        "size": "637 MB",
        "ram_required": 2,
        "speed": "Instant",
        "quality": "Good",
        "best_for": "Very slow / old PCs (under 4GB RAM)",
        "tier": 1,
    },
    "phi3": {
        "label": "Phi-3 (Default)",
        "size": "2.2 GB",
        "ram_required": 4,
        "speed": "Fast",
        "quality": "Very Good",
        "best_for": "Most laptops and PCs (4-8GB RAM)",
        "tier": 2,
    },
    "mistral": {
        "label": "Mistral 7B",
        "size": "4.1 GB",
        "ram_required": 8,
        "speed": "Moderate",
        "quality": "Excellent",
        "best_for": "Mid-range PC (8-16GB RAM)",
        "tier": 3,
    },
    "llama3": {
        "label": "LLaMA 3 8B",
        "size": "4.7 GB",
        "ram_required": 8,
        "speed": "Moderate",
        "quality": "Best",
        "best_for": "Fast PC (16GB+ RAM)",
        "tier": 4,
    },
}

def get_hardware_info() -> dict:
    try:
        ram_gb = round(psutil.virtual_memory().total / (1024 ** 3))
        cpu_cores = psutil.cpu_count(logical=False) or psutil.cpu_count()
        cpu_freq = psutil.cpu_freq()
        cpu_mhz = round(cpu_freq.current) if cpu_freq else 0
        return {"ram_gb": ram_gb, "cpu_cores": cpu_cores, "cpu_mhz": cpu_mhz, "platform": platform.processor() or platform.machine()}
    except Exception:
        return {"ram_gb": 4, "cpu_cores": 2, "cpu_mhz": 0, "platform": "unknown"}

def suggest_model(ram_gb: int) -> str:
    if ram_gb >= 16:
        return "llama3"
    elif ram_gb >= 8:
        return "mistral"
    elif ram_gb >= 4:
        return "phi3"
    else:
        return "tinyllama"

def get_installed_models() -> list:
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        if r.ok:
            models = r.json().get("models", [])
            return [m["name"].split(":")[0] for m in models]
    except Exception:
        pass
    return []


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def safe_str(val, default="AI analysis completed.") -> str:
    if val is None:
        return default
    return str(val).strip() or default

def safe_int(val, default=5, lo=1, hi=10) -> int:
    try:
        return max(lo, min(hi, int(val)))
    except Exception:
        return default

def extract_json(raw: str) -> Optional[dict]:
    """Safely extract JSON from any AI response."""
    try:
        start = raw.find("{")
        end = raw.rfind("}") + 1
        if start >= 0 and end > start:
            return json.loads(raw[start:end])
    except Exception:
        pass
    return None

def parse_risk_from_text(raw: str) -> Optional[dict]:
    """
    Parse risk score, priority and explanation from plain English AI response.
    Handles responses like:
      Risk Score: Low / Medium / High / 7
      Priority: P1 / High / Critical
      Explanation: ...
    """
    import re
    raw_lower = raw.lower()

    # ── Extract risk score ──
    score = None
    # Try numeric first: "Risk Score: 7" or "score: 8/10"
    m = re.search(r'risk.?score[:\s]+(\d+)', raw_lower)
    if m:
        score = int(m.group(1))
    else:
        # Try text-based score
        if any(w in raw_lower for w in ["risk score: high", "risk: high", "high risk"]):
            score = 8
        elif any(w in raw_lower for w in ["risk score: critical", "critical risk"]):
            score = 9
        elif any(w in raw_lower for w in ["risk score: medium", "risk score: moderate", "moderate risk", "medium risk"]):
            score = 6
        elif any(w in raw_lower for w in ["risk score: low", "low risk"]):
            score = 3

    # ── Extract priority ──
    priority = None
    m = re.search(r'priority[:\s]+(p[123]|high|medium|low|critical|moderate)', raw_lower)
    if m:
        val = m.group(1)
        priority_map = {"p1": "P1", "p2": "P2", "p3": "P3",
                        "critical": "P1", "high": "P1",
                        "medium": "P2", "moderate": "P2",
                        "low": "P3"}
        priority = priority_map.get(val, None)

    # ── Extract explanation ──
    explanation = None
    for pattern in [
        r'explanation[:\s]+(.+)',
        r'risk explanation[:\s]+(.+)',
        r'risk explanation title[:\s]+(.+)',
    ]:
        m = re.search(pattern, raw_lower)
        if m:
            explanation = m.group(1).strip().capitalize()
            break

    # If no explanation found, grab first meaningful sentence
    if not explanation:
        sentences = [s.strip() for s in raw.split(".") if len(s.strip()) > 20]
        if sentences:
            explanation = sentences[0][:200]

    # ── Derive missing fields ──
    if score is None and priority:
        score = {"P1": 8, "P2": 6, "P3": 3}[priority]
    if priority is None and score is not None:
        priority = "P1" if score >= 8 else "P2" if score >= 5 else "P3"

    if score is not None and priority is not None and explanation:
        return {
            "risk_score": max(1, min(10, score)),
            "priority": priority,
            "explanation": safe_str(explanation),
        }
    return None


def parse_automation_from_text(raw: str) -> Optional[dict]:
    """
    Parse automation suitability from plain English AI response.
    """
    import re
    raw_lower = raw.lower()

    # ── Extract suitability ──
    suitability = None
    if any(w in raw_lower for w in ["not suitable", "not automatable", "cannot be automated", "should not be automated"]):
        suitability = "Not Suitable"
    elif any(w in raw_lower for w in ["partially automatable", "partial automation", "partially suitable", "can be partially"]):
        suitability = "Partial"
    elif any(w in raw_lower for w in ["fully automatable", "automatable", "can be automated", "suitable for automation", "good candidate"]):
        suitability = "Automatable"

    # ── Extract confidence ──
    confidence = None
    m = re.search(r'confidence[:\s]+(\d+)', raw_lower)
    if m:
        confidence = int(m.group(1))
    else:
        if suitability == "Automatable":
            confidence = 80
        elif suitability == "Partial":
            confidence = 55
        elif suitability == "Not Suitable":
            confidence = 15

    # ── Extract explanation ──
    explanation = None
    for pattern in [r'explanation[:\s]+(.+)', r'reason[:\s]+(.+)']:
        m = re.search(pattern, raw_lower)
        if m:
            explanation = m.group(1).strip().capitalize()
            break
    if not explanation:
        sentences = [s.strip() for s in raw.split(".") if len(s.strip()) > 20]
        if sentences:
            explanation = sentences[0][:200]

    if suitability and confidence is not None and explanation:
        return {
            "suitability": suitability,
            "confidence": max(0, min(100, confidence)),
            "explanation": safe_str(explanation),
        }
    return None


# ─────────────────────────────────────────────
# LAYER 1 — Ollama AI (handles ANY test case)
# ─────────────────────────────────────────────
def call_ollama(prompt: str, model: str = None) -> str:
    use_model = model or OLLAMA_MODEL_DEFAULT
    response = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={"model": use_model, "prompt": prompt, "stream": False},
        timeout=120,
    )
    response.raise_for_status()
    return response.json().get("response", "")


def ollama_analyze_risk(title: str, description: str, steps: str, severity: str, model: str = None) -> Optional[dict]:
    """Call Ollama — parses both JSON and plain English responses."""
    prompt = f"""You are a QA risk analyst. Analyze this test case.

Title: {title}
Description: {description}
Severity: {severity}

Reply with ONLY these 3 lines (no extra text):
Risk Score: <number 1-10>
Priority: <P1 or P2 or P3>
Explanation: <one sentence>

Where P1=score 8-10, P2=score 5-7, P3=score 1-4"""

    try:
        raw = call_ollama(prompt, model)
        logger.info(f"Ollama raw response for {title[:30]}: {raw[:120]}")

        # Try JSON first
        parsed = extract_json(raw)
        if parsed:
            score = safe_int(parsed.get("risk_score") or parsed.get("score") or 5, default=5, lo=1, hi=10)
            priority = str(parsed.get("priority") or "").strip()
            if priority not in ("P1", "P2", "P3"):
                priority = "P1" if score >= 8 else "P2" if score >= 5 else "P3"
            explanation = safe_str(
                parsed.get("explanation") or parsed.get("reason") or
                parsed.get("rationale") or parsed.get("analysis")
            )
            logger.info(f"Ollama JSON risk: {title[:30]} → {priority} ({score})")
            return {"risk_score": score, "priority": priority, "explanation": explanation, "source": "ollama"}

        # Try plain English parsing (e.g. "Risk Score: Low\nPriority: Medium\nExplanation: ...")
        parsed_text = parse_risk_from_text(raw)
        if parsed_text:
            logger.info(f"Ollama text risk: {title[:30]} → {parsed_text['priority']} ({parsed_text['risk_score']})")
            return {**parsed_text, "source": "ollama"}

    except Exception as e:
        logger.warning(f"Ollama risk failed for '{title[:40]}': {e}")
    return None


def ollama_analyze_automation(title: str, description: str, steps: str, model: str = None) -> Optional[dict]:
    """Call Ollama — parses both JSON and plain English automation responses."""
    prompt = f"""You are a QA automation expert. Analyze this test case.

Title: {title}
Description: {description}
Steps: {steps}

Reply with ONLY these 3 lines (no extra text):
Suitability: <Automatable or Partial or Not Suitable>
Confidence: <number 0-100>
Explanation: <one sentence>

Rules:
- Automatable: stable, repeatable, no OTP/Captcha/biometric/physical hardware
- Partial: mostly automatable but has some dynamic or external elements
- Not Suitable: needs human, OTP, Captcha, biometric, voice, physical interaction"""

    try:
        raw = call_ollama(prompt, model)
        logger.info(f"Ollama auto raw for {title[:30]}: {raw[:120]}")

        # Try JSON first
        parsed = extract_json(raw)
        if parsed:
            suitability = str(parsed.get("suitability") or parsed.get("result") or "").strip()
            if suitability not in ("Automatable", "Partial", "Not Suitable"):
                suitability = "Partial"
            confidence = safe_int(parsed.get("confidence") or parsed.get("score") or 50, default=50, lo=0, hi=100)
            explanation = safe_str(parsed.get("explanation") or parsed.get("reason") or parsed.get("rationale"))
            logger.info(f"Ollama JSON auto: {title[:30]} → {suitability} ({confidence}%)")
            return {"suitability": suitability, "confidence": confidence, "explanation": explanation, "source": "ollama"}

        # Try plain English parsing
        parsed_text = parse_automation_from_text(raw)
        if parsed_text:
            logger.info(f"Ollama text auto: {title[:30]} → {parsed_text['suitability']} ({parsed_text['confidence']}%)")
            return {**parsed_text, "source": "ollama"}

    except Exception as e:
        logger.warning(f"Ollama automation failed for '{title[:40]}': {e}")
    return None


# ─────────────────────────────────────────────
# LAYER 2 — JSON-driven Keyword Heuristic
# Edit heuristic_rules.json to add/remove keywords.
# Hot-reload: rules refresh automatically when the JSON file changes.
# No server restart needed after editing the JSON.
# ─────────────────────────────────────────────

import pathlib as _pathlib

_RULES_JSON_PATH = _pathlib.Path(__file__).parent / "heuristic_rules.json"

# ── Internal cache ──────────────────────────────────────────────────────────
# Stores (mtime, RISK_RULES, AUTOMATION_RULES) so we only re-parse when the
# file actually changes on disk.
_rules_cache: dict = {"mtime": None, "risk": [], "auto": []}


def _load_rules() -> tuple:
    """
    Return (RISK_RULES, AUTOMATION_RULES) tuples from heuristic_rules.json.

    HOW HOT-RELOAD WORKS:
      - On every heuristic call we check the file's mtime (1 os.stat call ~1µs).
      - If mtime is unchanged we return the cached tuple immediately — zero I/O.
      - If the file changed we re-parse JSON and rebuild the in-memory tuples.
      - Thread-safe: Python's GIL protects the dict assignment on CPython.

    JSON SCHEMA expected:
      {
        "risk_rules":       [{"keywords": [...], "score_boost": int, "label": str}, ...],
        "automation_rules": [{"keywords": [...], "suitability": str, "confidence": int, "reason": str}, ...]
      }
    """
    try:
        mtime = _RULES_JSON_PATH.stat().st_mtime
    except FileNotFoundError:
        logger.error(f"[Rules] heuristic_rules.json not found at {_RULES_JSON_PATH} — using cached rules")
        return _rules_cache["risk"], _rules_cache["auto"]

    # Cache hit — file unchanged
    if mtime == _rules_cache["mtime"]:
        return _rules_cache["risk"], _rules_cache["auto"]

    # Cache miss — file changed or first load
    try:
        raw = json.loads(_RULES_JSON_PATH.read_text(encoding="utf-8"))

        risk_rules = [
            (r["keywords"], r["score_boost"], r["label"])
            for r in raw.get("risk_rules", [])
        ]
        auto_rules = [
            (r["keywords"], r["suitability"], r["confidence"], r["reason"])
            for r in raw.get("automation_rules", [])
        ]

        _rules_cache["mtime"] = mtime
        _rules_cache["risk"]  = risk_rules
        _rules_cache["auto"]  = auto_rules

        total_risk = sum(len(r[0]) for r in risk_rules)
        total_auto = sum(len(r[0]) for r in auto_rules)
        logger.info(
            f"[Rules] Loaded heuristic_rules.json — "
            f"{len(risk_rules)} risk groups ({total_risk} keywords), "
            f"{len(auto_rules)} automation groups ({total_auto} keywords)"
        )
        return risk_rules, auto_rules

    except Exception as e:
        logger.error(f"[Rules] Failed to parse heuristic_rules.json: {e} — using cached rules")
        return _rules_cache["risk"], _rules_cache["auto"]


# Pre-load at startup so the first request has no cold-start delay
try:
    _load_rules()
except Exception:
    pass

def heuristic_risk(title: str, description: str, steps: str, severity: str) -> dict:
    """
    JSON-driven keyword heuristic for risk scoring.
    Rules are loaded from heuristic_rules.json — edit the JSON to add keywords,
    changes are picked up automatically without restarting the server.
    """
    risk_rules, _ = _load_rules()

    text = (title + " " + description + " " + steps).lower()
    sev  = severity.lower()
    score = 3
    reason_parts = []

    # Severity scoring
    if sev in ("critical", "blocker", "p0"):
        score += 4
        reason_parts.append("critical severity")
    elif sev in ("high", "major", "p1"):
        score += 2
        reason_parts.append("high severity")
    elif sev in ("medium", "moderate", "p2"):
        score += 1

    # Keyword scoring — uses hot-reloaded rules from JSON
    matched = False
    for keywords, boost, label in risk_rules:
        if any(k in text for k in keywords):
            score += boost
            reason_parts.append(label)
            matched = True
            break

    if not matched:
        reason_parts.append("general functionality area")

    score = max(1, min(10, score))
    priority = "P1" if score >= 8 else "P2" if score >= 5 else "P3"
    explanation = f"Risk assessment based on {', '.join(reason_parts)}. Score: {score}/10."
    return {"risk_score": score, "priority": priority, "explanation": explanation, "source": "heuristic"}


def heuristic_automation(title: str, description: str, steps: str) -> dict:
    """
    JSON-driven keyword heuristic for automation suitability.
    Rules are loaded from heuristic_rules.json — edit the JSON to add keywords,
    changes are picked up automatically without restarting the server.
    """
    _, auto_rules = _load_rules()

    text = (title + " " + description + " " + steps).lower()

    for keywords, suitability, confidence, reason in auto_rules:
        if any(k in text for k in keywords):
            return {"suitability": suitability, "confidence": confidence, "explanation": reason, "source": "heuristic"}

    return {
        "suitability": "Partial",
        "confidence": 50,
        "explanation": "Unable to determine automation suitability from content — manual review recommended.",
        "source": "heuristic"
    }


# ─────────────────────────────────────────────
# LAYER 3 — Smart Dispatcher
# Tries Ollama first → falls back to heuristic
# ─────────────────────────────────────────────
def offline_analyze_risk(row: dict, model: str = None) -> dict:
    title = safe_str(row.get("title") or row.get("Title") or row.get("Test Case Title"), "Untitled")
    description = safe_str(row.get("description") or row.get("Description") or row.get("Test Description"), "")
    steps = safe_str(row.get("steps") or row.get("Steps") or row.get("Test Steps"), "")
    severity = safe_str(row.get("severity") or row.get("Severity"), "")

    # Try Ollama AI first — handles any test case
    result = ollama_analyze_risk(title, description, steps, severity, model)
    if result:
        result["mode"] = "offline"
        return result

    # Fallback to extended heuristic
    logger.info(f"Using heuristic fallback for: {title[:40]}")
    result = heuristic_risk(title, description, steps, severity)
    result["mode"] = "offline"
    result["model_used"] = "heuristic"
    return result


def offline_analyze_automation(row: dict, model: str = None) -> dict:
    title = safe_str(row.get("title") or row.get("Title") or row.get("Test Case Title"), "Untitled")
    description = safe_str(row.get("description") or row.get("Description") or row.get("Test Description"), "")
    steps = safe_str(row.get("steps") or row.get("Steps") or row.get("Test Steps"), "")

    # Try Ollama AI first — handles any test case
    result = ollama_analyze_automation(title, description, steps, model)
    if result:
        result["mode"] = "offline"
        return result

    # Fallback to extended heuristic
    logger.info(f"Using heuristic fallback for: {title[:40]}")
    result = heuristic_automation(title, description, steps)
    result["mode"] = "offline"
    return result


# ─────────────────────────────────────────────
# ONLINE MODE — OpenAI (SDK, gpt-4o-mini, strict JSON)
# IMPROVEMENTS: 1 (SDK + model), 4 (strict JSON), 6 (helpers)
# ─────────────────────────────────────────────

# IMPROVEMENT 6: shared field extractors
def _extract_risk_fields(row: dict):
    title = safe_str(row.get("title") or row.get("Title") or row.get("Test Case Title"), "Untitled")
    description = safe_str(row.get("description") or row.get("Description") or row.get("Test Description"), "")
    steps = safe_str(row.get("steps") or row.get("Steps") or row.get("Test Steps"), "")
    severity = safe_str(row.get("severity") or row.get("Severity"), "")
    return title, description, steps, severity

def _extract_auto_fields(row: dict):
    title = safe_str(row.get("title") or row.get("Title") or row.get("Test Case Title"), "Untitled")
    description = safe_str(row.get("description") or row.get("Description") or row.get("Test Description"), "")
    steps = safe_str(row.get("steps") or row.get("Steps") or row.get("Test Steps"), "")
    return title, description, steps

# IMPROVEMENT 1: Official OpenAI SDK, gpt-4o-mini, strict JSON
def call_openai_sdk(prompt: str, api_key: str, system: str) -> str:
    client = OpenAI(api_key=api_key, timeout=30.0, max_retries=0)
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        response_format={"type": "json_object"},
        temperature=0.1,
        max_tokens=200,
    )
    return response.choices[0].message.content


def openai_analyze_risk(row: dict, api_key: str) -> dict:
    title, description, steps, severity = _extract_risk_fields(row)
    system = (
        "You are a senior QA risk analyst. Respond with valid JSON only. "
        "Keys: risk_score (int 1-10), priority (P1|P2|P3), explanation (string). "
        "P1=8-10 critical, P2=5-7 moderate, P3=1-4 low."
    )
    prompt = f"Title: {title}\nDescription: {description}\nSteps: {steps}\nSeverity: {severity}"
    try:
        raw = call_openai_sdk(prompt, api_key, system)
        parsed = json.loads(raw)
        score = safe_int(parsed.get("risk_score") or 5, default=5, lo=1, hi=10)
        priority = str(parsed.get("priority") or "").strip()
        if priority not in ("P1", "P2", "P3"):
            priority = "P1" if score >= 8 else "P2" if score >= 5 else "P3"
        explanation = safe_str(parsed.get("explanation") or parsed.get("reason"))
        return {"risk_score": score, "priority": priority, "explanation": explanation, "mode": "online", "source": "openai"}
    except RateLimitError:
        logger.warning(f"OpenAI rate limited — falling back to Ollama for: {title[:40]}")
        result = ollama_analyze_risk(title, description, steps, severity) or heuristic_risk(title, description, steps, severity)
        result["mode"] = "online-fallback"
        return result
    except AuthenticationError:
        raise HTTPException(status_code=401, detail="Invalid OpenAI API key.")
    except (APIError, Exception) as e:
        logger.warning(f"OpenAI error — falling back to Ollama: {e}")
        result = ollama_analyze_risk(title, description, steps, severity) or heuristic_risk(title, description, steps, severity)
        result["mode"] = "online-fallback"
        return result


def openai_analyze_automation(row: dict, api_key: str) -> dict:
    title, description, steps = _extract_auto_fields(row)
    system = (
        "You are a senior QA automation architect. Respond with valid JSON only. "
        "Keys: suitability (Automatable|Partial|Not Suitable), confidence (int 0-100), explanation (string). "
        "Automatable=stable repeatable, Partial=some dynamic elements, Not Suitable=needs human/OTP/biometric."
    )
    prompt = f"Title: {title}\nDescription: {description}\nSteps: {steps}"
    try:
        raw = call_openai_sdk(prompt, api_key, system)
        parsed = json.loads(raw)
        suitability = str(parsed.get("suitability") or "").strip()
        if suitability not in ("Automatable", "Partial", "Not Suitable"):
            suitability = "Partial"
        confidence = safe_int(parsed.get("confidence") or 50, default=50, lo=0, hi=100)
        explanation = safe_str(parsed.get("explanation") or parsed.get("reason"))
        return {"suitability": suitability, "confidence": confidence, "explanation": explanation, "mode": "online", "source": "openai"}
    except RateLimitError:
        logger.warning(f"OpenAI rate limited — falling back to Ollama for: {title[:40]}")
        result = ollama_analyze_automation(title, description, steps) or heuristic_automation(title, description, steps)
        result["mode"] = "online-fallback"
        return result
    except AuthenticationError:
        raise HTTPException(status_code=401, detail="Invalid OpenAI API key.")
    except (APIError, Exception) as e:
        raise HTTPException(status_code=502, detail=f"OpenAI error: {str(e)}")


# ─────────────────────────────────────────────
# GEMINI MODE — Google Gemini AI
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# GEMINI — Model cache, rate limiter, safe helpers
# ─────────────────────────────────────────────

# Fix 3 + 4: module-level cache — list_models() called ONCE per key, never per request
_gemini_model_cache: dict = {}   # key_fingerprint → model_name string
_gemini_list_lock = threading.Lock()

# Fix 1 + 2: sliding-window rate limiter — free tier = 5 RPM max
_gemini_rpm_limit  = 5       # max requests per 60 s
_gemini_rpm_window = 60.0    # seconds
_gemini_timestamps: dict = {}   # key_fingerprint → collections.deque of monotonic timestamps
_gemini_rate_lock = threading.Lock()


def _gemini_rate_wait(key_fingerprint: str):
    """
    Block the calling thread until it is safe to send another Gemini request.
    Implements a sliding-window rate limiter capped at _gemini_rpm_limit RPM.
    Safe to call from ThreadPoolExecutor workers — time.sleep() is intentional.
    """
    with _gemini_rate_lock:
        if key_fingerprint not in _gemini_timestamps:
            _gemini_timestamps[key_fingerprint] = collections.deque()
        timestamps = _gemini_timestamps[key_fingerprint]

        now = time.monotonic()
        # Drop timestamps older than the window
        while timestamps and now - timestamps[0] >= _gemini_rpm_window:
            timestamps.popleft()

        if len(timestamps) >= _gemini_rpm_limit:
            # Must wait until the oldest request exits the window
            wait = _gemini_rpm_window - (now - timestamps[0]) + 0.1
            logger.info(f"[Gemini] Rate limit — waiting {wait:.1f}s (5 RPM free tier)")
            time.sleep(wait)
            # Re-clean after sleep
            now = time.monotonic()
            while timestamps and now - timestamps[0] >= _gemini_rpm_window:
                timestamps.popleft()

        timestamps.append(time.monotonic())


def _discover_gemini_model(client) -> str:
    """
    Fix 3: Call list_models() ONCE and pick the best generateContent model.
    Sorted by: flash > pro > other, 2.0 > 1.5 > 1.0, -latest > plain.
    """
    try:
        all_models = list(client.models.list())
    except Exception as e:
        raise RuntimeError(f"[Gemini] Cannot list models: {e}")

    # Filter to models that support generateContent
    supported = []
    for m in all_models:
        actions = (
            getattr(m, "supported_actions", None) or
            getattr(m, "supported_generation_methods", None) or
            []
        )
        if "generateContent" in actions:
            supported.append(m)

    # If SDK doesn't expose actions, use full list
    if not supported:
        supported = all_models

    names = [getattr(m, "name", None) or str(m) for m in supported]
    logger.info(f"[Gemini] Models available for generateContent: {names}")

    def priority(name: str) -> tuple:
        n = name.lower()
        tier    = 0 if "flash" in n else 1 if "pro" in n else 2
        version = 0 if "2.0" in n else 1 if "1.5" in n else 2 if "1.0" in n else 3
        latest  = 0 if "latest" in n else 1
        return (tier, version, latest)

    names.sort(key=priority)
    if not names:
        raise RuntimeError("[Gemini] No generateContent-capable models found for this API key.")

    logger.info(f"[Gemini] Selected model: {names[0]}")
    return names[0]


def _get_gemini_model(client, key_fingerprint: str) -> str:
    """
    Fix 4: Return cached model name, discovering only on first call per key.
    Thread-safe — multiple parallel workers share the same cache safely.
    """
    with _gemini_list_lock:
        if key_fingerprint not in _gemini_model_cache:
            _gemini_model_cache[key_fingerprint] = _discover_gemini_model(client)
        return _gemini_model_cache[key_fingerprint]


def _safe_parse_gemini_json(raw: str, context: str) -> dict:
    """
    Fix 5: Safe JSON parser — returns None if response is empty or unparseable.
    Never raises — caller decides what to do on None.
    """
    if not raw or not raw.strip():
        logger.warning(f"[Gemini] Empty response for {context} — will fallback")
        return None
    try:
        # Strip accidental markdown fences Gemini sometimes adds despite mime type
        cleaned = raw.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        return json.loads(cleaned)
    except json.JSONDecodeError as e:
        logger.warning(f"[Gemini] JSON parse failed for {context}: {e} | raw={raw[:80]!r}")
        return None


def call_gemini(prompt: str, api_key: str) -> str:
    """
    Fix 1-4: Rate-limited, model-cached Gemini call.
    - Waits automatically if 5 RPM free-tier limit is close
    - Discovers and caches model on first call per API key
    - Returns raw response text (may be empty — callers use _safe_parse_gemini_json)
    - On 404, clears model cache and retries discovery once
    """
    key_fingerprint = api_key[-8:]
    client = genai.Client(api_key=api_key)

    # Fix 1+2: enforce 5 RPM before every call
    _gemini_rate_wait(key_fingerprint)

    # Fix 4: get cached model (discovers on first call)
    model_name = _get_gemini_model(client, key_fingerprint)

    def _call(model: str) -> str:
        response = client.models.generate_content(
            model=model,
            contents=prompt,
            config=genai_types.GenerateContentConfig(
                temperature=0.1,
                max_output_tokens=300,
                response_mime_type="application/json",
            ),
        )
        return response.text or ""

    try:
        return _call(model_name)
    except Exception as e:
        err = str(e)
        if "404" in err or "NOT_FOUND" in err:
            # Cached model gone — rediscover once
            logger.warning(f"[Gemini] Model '{model_name}' gone — rediscovering...")
            with _gemini_list_lock:
                _gemini_model_cache.pop(key_fingerprint, None)
            model_name = _get_gemini_model(client, key_fingerprint)
            return _call(model_name)
        raise


def _classify_gemini_error(e: Exception) -> str:
    """
    Precisely classify a Gemini exception so each error type gets
    the correct treatment instead of everything becoming a 401.

    Returns one of four string labels:
      "invalid_key"  — key is genuinely wrong/revoked   → raise HTTP 401
      "quota"        — free-tier RPM or daily cap hit    → fallback silently
      "network"      — timeout, DNS, connection refused  → fallback silently
      "fallback"     — everything else (empty resp, etc) → fallback silently

    Root cause of the original bug:
      The old check matched "permission" inside the word "RESOURCE_EXHAUSTED"
      and "invalid" inside "invalid argument" (a quota-related Gemini message),
      so quota errors were wrongly classified as auth failures and returned 401.
    """
    err = str(e).lower()
    original = str(e)   # keep original case for status-code checks

    # ── Genuine auth failures ────────────────────────────────────────────────
    # Only raise 401 when Google explicitly says the key itself is bad.
    # Match on the exact status code 401 or Google's specific auth messages.
    # Do NOT match on "invalid" or "permission" alone — those appear in quota
    # and other non-auth error messages too.
    AUTH_SIGNALS = [
        "api_key_invalid",           # Google error code
        "api key not valid",         # Google human-readable message
        "invalid api key",           # variation
        "unauthenticated",           # gRPC / HTTP status word
        "401",                       # raw HTTP status in message
    ]
    if any(sig in err for sig in AUTH_SIGNALS):
        return "invalid_key"

    # ── Quota / rate-limit errors ────────────────────────────────────────────
    # 429 RESOURCE_EXHAUSTED covers both RPM and daily quota.
    # "quota_exceeded" and "rateLimitExceeded" are Google error codes.
    QUOTA_SIGNALS = [
        "429",
        "resource_exhausted",        # gRPC status
        "quota_exceeded",            # Google error code
        "ratelimitexceeded",         # variation
        "rate limit",
        "quota",
        "too many requests",
        "daily limit",
        "per minute",
    ]
    if any(sig in err for sig in QUOTA_SIGNALS):
        return "quota"

    # ── Network / infrastructure errors ─────────────────────────────────────
    NETWORK_SIGNALS = [
        "timeout",
        "timed out",
        "connection",
        "network",
        "dns",
        "socket",
        "unreachable",
        "name or service not known",
        "ssl",
        "certificate",
        "503",
        "502",
        "504",
    ]
    if any(sig in err for sig in NETWORK_SIGNALS):
        return "network"

    # ── Everything else (empty response, JSON parse fail, model gone, etc) ───
    return "fallback"


def gemini_analyze_risk(row: dict, api_key: str) -> dict:
    """
    Risk analysis via Gemini with precise error classification.
    Only raises HTTP 401 when the key is genuinely invalid.
    All other errors (quota, network, empty response) fall back
    silently to Ollama → Heuristic.
    """
    title, description, steps, severity = _extract_risk_fields(row)
    prompt = (
        "You are a senior QA risk analyst. Respond with valid JSON only.\n"
        f"Title: {title}\nDescription: {description}\nSteps: {steps}\nSeverity: {severity}\n\n"
        "Return exactly this JSON structure:\n"
        '{"risk_score": 7, "priority": "P2", "explanation": "one professional sentence"}\n'
        "Rules: risk_score integer 1-10. priority: P1 (score 8-10), P2 (5-7), P3 (1-4)."
    )
    try:
        raw    = call_gemini(prompt, api_key)
        parsed = _safe_parse_gemini_json(raw, title[:30])

        if parsed is None:
            raise ValueError("Empty or unparseable Gemini response")

        score = safe_int(parsed.get("risk_score") or 5, default=5, lo=1, hi=10)
        priority = str(parsed.get("priority") or "").strip()
        if priority not in ("P1", "P2", "P3"):
            priority = "P1" if score >= 8 else "P2" if score >= 5 else "P3"
        explanation = safe_str(parsed.get("explanation") or parsed.get("reason"))
        logger.info(f"[Gemini] Risk: {title[:30]} → {priority} (score={score})")
        return {"risk_score": score, "priority": priority, "explanation": explanation,
                "mode": "gemini", "source": "gemini"}

    except Exception as e:
        error_type = _classify_gemini_error(e)

        if error_type == "invalid_key":
            # Only case that surfaces to the frontend as 401
            logger.warning(f"[Gemini] Invalid API key detected for '{title[:40]}'")
            raise HTTPException(status_code=401, detail="Invalid Gemini API key.")

        elif error_type == "quota":
            logger.warning(f"[Gemini] Quota/rate-limit for '{title[:40]}' — falling back to Ollama")

        elif error_type == "network":
            logger.warning(f"[Gemini] Network error for '{title[:40]}': {e} — falling back to Ollama")

        else:
            # empty response, JSON parse fail, unexpected error — all fall back
            logger.warning(f"[Gemini] Fallback for '{title[:40]}': {e}")

        result = ollama_analyze_risk(title, description, steps, severity) or \
                 heuristic_risk(title, description, steps, severity)
        result["mode"] = "gemini-fallback"
        return result


def gemini_analyze_automation(row: dict, api_key: str) -> dict:
    """
    Automation analysis via Gemini with precise error classification.
    Only raises HTTP 401 when the key is genuinely invalid.
    All other errors fall back silently to Ollama → Heuristic.
    """
    title, description, steps = _extract_auto_fields(row)
    prompt = (
        "You are a senior QA automation architect. Respond with valid JSON only.\n"
        f"Title: {title}\nDescription: {description}\nSteps: {steps}\n\n"
        "Return exactly this JSON structure:\n"
        '{"suitability": "Automatable", "confidence": 85, "explanation": "one professional sentence"}\n'
        "Rules: suitability must be exactly one of: Automatable, Partial, Not Suitable.\n"
        "Automatable=stable repeatable, no OTP/Captcha.\n"
        "Partial=some dynamic/external dependencies.\n"
        "Not Suitable=needs human judgment, OTP, Captcha, biometric, or physical hardware."
    )
    try:
        raw    = call_gemini(prompt, api_key)
        parsed = _safe_parse_gemini_json(raw, title[:30])

        if parsed is None:
            raise ValueError("Empty or unparseable Gemini response")

        suitability = str(parsed.get("suitability") or "").strip()
        if suitability not in ("Automatable", "Partial", "Not Suitable"):
            suitability = "Partial"
        confidence  = safe_int(parsed.get("confidence") or 50, default=50, lo=0, hi=100)
        explanation = safe_str(parsed.get("explanation") or parsed.get("reason"))
        logger.info(f"[Gemini] Automation: {title[:30]} → {suitability} ({confidence}%)")
        return {"suitability": suitability, "confidence": confidence, "explanation": explanation,
                "mode": "gemini", "source": "gemini"}

    except Exception as e:
        error_type = _classify_gemini_error(e)

        if error_type == "invalid_key":
            logger.warning(f"[Gemini] Invalid API key detected for '{title[:40]}'")
            raise HTTPException(status_code=401, detail="Invalid Gemini API key.")

        elif error_type == "quota":
            logger.warning(f"[Gemini] Quota/rate-limit for '{title[:40]}' — falling back to Ollama")

        elif error_type == "network":
            logger.warning(f"[Gemini] Network error for '{title[:40]}': {e} — falling back to Ollama")

        else:
            logger.warning(f"[Gemini] Fallback for '{title[:40]}': {e}")

        result = ollama_analyze_automation(title, description, steps) or \
                 heuristic_automation(title, description, steps)
        result["mode"] = "gemini-fallback"
        return result


# ─────────────────────────────────────────────
# Excel generators
# ─────────────────────────────────────────────

# ── AI source display helpers ─────────────────────────────────────────────────

# Colour palette per AI engine — background fill for the AI Source cell
_AI_SOURCE_FILLS = {
    "gemini":          PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),  # soft green
    "gemini-fallback": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # soft yellow
    "openai":          PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),  # soft blue
    "openai-fallback": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),  # soft yellow
    "offline":         PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),  # soft purple
    "heuristic":       PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # soft lavender
}
_AI_SOURCE_FONT_COLORS = {
    "gemini":          "1B5E20",   # dark green
    "gemini-fallback": "F57F17",   # dark amber
    "openai":          "0D47A1",   # dark blue
    "openai-fallback": "F57F17",   # dark amber
    "offline":         "4A148C",   # dark purple
    "heuristic":       "6A1B9A",   # deep purple
}
_AI_SOURCE_LABELS = {
    "gemini":          "🟢 Google Gemini",
    "gemini-fallback": "⚠️ Gemini→Fallback",
    "openai":          "🔵 OpenAI GPT",
    "openai-fallback": "⚠️ OpenAI→Fallback",
    "offline":         "🔌 Local Ollama",
    "heuristic":       "📋 Keyword Rules",
}

def _ai_source_label(raw_mode: str) -> str:
    """Convert internal mode string to a human-readable label with icon."""
    key = str(raw_mode).lower().strip()
    return _AI_SOURCE_LABELS.get(key, f"🤖 {raw_mode.upper()}")

def _apply_ai_source_cell(cell, raw_mode: str):
    """Apply colour fill + coloured bold text to an AI Source cell."""
    key = str(raw_mode).lower().strip()
    fill  = _AI_SOURCE_FILLS.get(key,      PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"))
    fcolor = _AI_SOURCE_FONT_COLORS.get(key, "424242")
    cell.fill = fill
    cell.font = Font(bold=True, color=fcolor, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def _engine_banner(mode: str) -> str:
    """One-line banner string describing which AI engine produced this report."""
    labels = {
        "gemini":  "Google Gemini AI",
        "openai":  "OpenAI GPT-4o-mini",
        "offline": "Local Ollama (Offline AI)",
    }
    return labels.get(mode.lower(), mode.upper())

def _summary_engine_rows(ws, mode: str, data_start_row: int):
    """
    Write the AI Engine section at the bottom of a Summary sheet.
    Shows engine name, colour, and what fallback means.
    """
    import datetime
    engine = _engine_banner(mode)
    ws.append([])
    ws.append(["─── AI Engine Information ───", "", ""])
    ws.append(["Primary AI Engine", engine, ""])
    ws.append(["Fallback Engine",   "Local Ollama → Keyword Rules", ""])
    ws.append(["Report Generated",  datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), ""])
    ws.append([])
    ws.append(["AI Source Legend", "", ""])

    legend = [
        ("🟢 Google Gemini",    "E8F5E9", "1B5E20", "Result produced by Google Gemini AI"),
        ("🔵 OpenAI GPT",       "E3F2FD", "0D47A1", "Result produced by OpenAI GPT-4o-mini"),
        ("🔌 Local Ollama",     "EDE7F6", "4A148C", "Result produced by local Ollama model"),
        ("📋 Keyword Rules",    "F3E5F5", "6A1B9A", "Result from keyword heuristic engine"),
        ("⚠️ Gemini→Fallback",  "FFF9C4", "F57F17", "Gemini failed — result from fallback engine"),
        ("⚠️ OpenAI→Fallback",  "FFF9C4", "F57F17", "OpenAI failed — result from fallback engine"),
    ]
    for label, bg, fg, desc in legend:
        r = ws.max_row + 1
        ws.append([label, desc, ""])
        ws.cell(row=r, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws.cell(row=r, column=1).font = Font(bold=True, color=fg, size=10)
        ws.cell(row=r, column=2).font = Font(color="424242", size=10)


def create_regression_excel(df: pd.DataFrame, session_id: str, mode: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Regression Results"

    # ── Header row ──────────────────────────────────────────────────────────
    headers = ["#", "Test Case Title", "Description", "Risk Score", "Priority",
               "AI Risk Explanation", "Recommended", "AI Source"]
    ws.append(headers)

    hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hfill
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Data rows ────────────────────────────────────────────────────────────
    priority_colors = {"P1": "FFCCCC", "P2": "FFF2CC", "P3": "D9EAD3"}

    for i, row in df.iterrows():
        title       = safe_str(row.get("Title") or row.get("title") or row.get("Test Case Title"), f"TC-{i+1}")
        desc        = safe_str(row.get("Description") or row.get("description"), "")
        risk_score  = row.get("Risk Score", 5)
        priority    = str(row.get("Priority", "P2"))
        explanation = safe_str(row.get("AI Risk Explanation"), "")
        recommended = str(row.get("Recommended for Execution", "No"))
        raw_source  = str(row.get("AI Source") or row.get("AI Mode") or mode)
        label       = _ai_source_label(raw_source)

        ws.append([i + 1, title, desc, risk_score, priority, explanation, recommended, label])

        data_row = i + 2
        # Priority background on columns A–G
        color = priority_colors.get(priority, "FFFFFF")
        rfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, 8):
            c = ws.cell(row=data_row, column=col)
            c.fill = rfill
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # AI Source column (col 8) — its own colour per engine
        _apply_ai_source_cell(ws.cell(row=data_row, column=8), raw_source)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 22   # AI Source — wider for icon + label
    ws.row_dimensions[1].height = 35

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    p1    = len(df[df["Priority"] == "P1"])
    p2    = len(df[df["Priority"] == "P2"])
    p3    = len(df[df["Priority"] == "P3"])
    total = len(df)
    rec   = len(df[df["Recommended for Execution"] == "Yes"])

    ws2.append([f"AI QA Decision Intelligence — Regression Report", "", ""])
    ws2.merge_cells("A1:C1")
    ws2["A1"].font      = Font(bold=True, size=14, color="1E3A5F")
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.append([f"AI Engine: {_engine_banner(mode)}", "", ""])
    ws2.merge_cells("A2:C2")
    ws2["A2"].font      = Font(bold=True, size=11, color="2E75B6")
    ws2["A2"].alignment = Alignment(horizontal="center")

    ws2.append([])
    ws2.append(["Metric", "Count", "Percentage"])
    ws2.append(["Total Test Cases",          total, "100%"])
    ws2.append(["P1 — Critical",             p1,    f"{round(p1/total*100)}%" if total else "0%"])
    ws2.append(["P2 — Moderate",             p2,    f"{round(p2/total*100)}%" if total else "0%"])
    ws2.append(["P3 — Low Risk",             p3,    f"{round(p3/total*100)}%" if total else "0%"])
    ws2.append(["Recommended for Execution", rec,   f"{round(rec/total*100)}%" if total else "0%"])

    sfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col in range(1, 4):
        c = ws2.cell(row=4, column=col)
        c.fill = sfill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    # AI source breakdown — how many rows came from each engine
    ws2.append([])
    ws2.append(["AI Source Breakdown", "Count", "% of Total"])
    src_header_row = ws2.max_row
    src_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    for col in range(1, 4):
        c = ws2.cell(row=src_header_row, column=col)
        c.fill = src_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    source_col = "AI Source" if "AI Source" in df.columns else "AI Mode"
    if source_col in df.columns:
        for src_val, count in df[source_col].value_counts().items():
            pct = f"{round(count/total*100)}%" if total else "0%"
            label = _ai_source_label(str(src_val))
            r = ws2.max_row + 1
            ws2.append([label, count, pct])
            _apply_ai_source_cell(ws2.cell(row=r, column=1), str(src_val))
            ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            ws2.cell(row=r, column=3).alignment = Alignment(horizontal="center")

    _summary_engine_rows(ws2, mode, ws2.max_row)

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 32

    path = os.path.join(OUTPUT_DIR, f"regression_results_{session_id}.xlsx")
    wb.save(path)
    return path


def create_automation_excel(df: pd.DataFrame, session_id: str, mode: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Automation Analysis"

    # ── Header row ──────────────────────────────────────────────────────────
    headers = ["#", "Test Case Title", "Description", "Automation Suitability",
               "Confidence %", "AI Explanation", "AI Source"]
    ws.append(headers)

    hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = hfill
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Data rows ────────────────────────────────────────────────────────────
    suit_colors = {"Automatable": "D9EAD3", "Partial": "FFF2CC", "Not Suitable": "FFCCCC"}

    for i, row in df.iterrows():
        title       = safe_str(row.get("Title") or row.get("title") or row.get("Test Case Title"), f"TC-{i+1}")
        desc        = safe_str(row.get("Description") or row.get("description"), "")
        suitability = str(row.get("Automation Suitability", "Partial"))
        confidence  = row.get("Confidence %", 50)
        explanation = safe_str(row.get("AI Explanation"), "")
        raw_source  = str(row.get("AI Source") or row.get("AI Mode") or mode)
        label       = _ai_source_label(raw_source)

        ws.append([i + 1, title, desc, suitability, f"{confidence}%", explanation, label])

        data_row = i + 2
        # Suitability background on columns A–F
        color = suit_colors.get(suitability, "FFFFFF")
        rfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, 7):
            c = ws.cell(row=data_row, column=col)
            c.fill = rfill
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # AI Source column (col 7) — its own colour per engine
        _apply_ai_source_cell(ws.cell(row=data_row, column=7), raw_source)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 22   # AI Source
    ws.row_dimensions[1].height = 35

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total        = len(df)
    automatable  = len(df[df["Automation Suitability"] == "Automatable"])
    partial      = len(df[df["Automation Suitability"] == "Partial"])
    not_suitable = len(df[df["Automation Suitability"] == "Not Suitable"])

    ws2.append([f"AI QA Decision Intelligence — Automation Report", "", ""])
    ws2.merge_cells("A1:C1")
    ws2["A1"].font      = Font(bold=True, size=14, color="1E3A5F")
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.append([f"AI Engine: {_engine_banner(mode)}", "", ""])
    ws2.merge_cells("A2:C2")
    ws2["A2"].font      = Font(bold=True, size=11, color="2E75B6")
    ws2["A2"].alignment = Alignment(horizontal="center")

    ws2.append([])
    ws2.append(["Category", "Count", "Percentage"])
    ws2.append(["Total Analyzed",  total,        "100%"])
    ws2.append(["Automatable",     automatable,  f"{round(automatable/total*100)}%"  if total else "0%"])
    ws2.append(["Partial",         partial,      f"{round(partial/total*100)}%"      if total else "0%"])
    ws2.append(["Not Suitable",    not_suitable, f"{round(not_suitable/total*100)}%" if total else "0%"])

    sfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col in range(1, 4):
        c = ws2.cell(row=4, column=col)
        c.fill = sfill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    # AI source breakdown
    ws2.append([])
    ws2.append(["AI Source Breakdown", "Count", "% of Total"])
    src_header_row = ws2.max_row
    src_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    for col in range(1, 4):
        c = ws2.cell(row=src_header_row, column=col)
        c.fill = src_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    source_col = "AI Source" if "AI Source" in df.columns else "AI Mode"
    if source_col in df.columns:
        for src_val, count in df[source_col].value_counts().items():
            pct = f"{round(count/total*100)}%" if total else "0%"
            label = _ai_source_label(str(src_val))
            r = ws2.max_row + 1
            ws2.append([label, count, pct])
            _apply_ai_source_cell(ws2.cell(row=r, column=1), str(src_val))
            ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
            ws2.cell(row=r, column=3).alignment = Alignment(horizontal="center")

    _summary_engine_rows(ws2, mode, ws2.max_row)

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 32

    path = os.path.join(OUTPUT_DIR, f"automation_results_{session_id}.xlsx")
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# API Routes
# ─────────────────────────────────────────────
@app.get("/")
def root():
    return {"status": "AI QA Decision Intelligence Platform Running", "version": "1.0"}


@app.get("/health")
def health():
    # IMPROVEMENT 2 (Bug fix): was using undefined OLLAMA_MODEL, now uses OLLAMA_MODEL_DEFAULT
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        ollama_status = "connected" if r.ok else "error"
    except Exception:
        ollama_status = "not reachable"
    return {"api": "healthy", "ollama": ollama_status, "model": OLLAMA_MODEL_DEFAULT}


@app.get("/api/test")
def test():
    return {"status": "backend working", "output_dir": OUTPUT_DIR}


@app.get("/api/rate-stats")
def rate_stats():
    """Live monitoring — token bucket status, cache hit rate, queue size."""
    return ai_manager.get_stats()


@app.get("/api/ai-status")
def ai_status():
    """Returns real-time rate limit usage, cache stats and queue depth."""
    return ai_manager.get_status()


@app.post("/api/validate-key")
async def validate_key(data: dict):
    """
    Auto-detect key type from format and validate.
    OpenAI keys start with: sk-
    Gemini keys start with: AIza
    """
    api_key = data.get("api_key", "").strip()

    if not api_key:
        raise HTTPException(status_code=400, detail="API key is required.")

    # ── Auto-detect key type ──
    if api_key.startswith("sk-"):
        key_type = "openai"
    elif api_key.startswith("AIza"):
        key_type = "gemini"
    else:
        return {
            "valid": False,
            "key_type": "unknown",
            "message": "Cannot identify key type. OpenAI keys start with 'sk-', Gemini keys start with 'AIza'."
        }

    # ── Validate OpenAI ──
    if key_type == "openai":
        try:
            client = OpenAI(api_key=api_key, timeout=10.0, max_retries=0)
            client.models.list()
            return {"valid": True, "key_type": "openai", "message": "OpenAI API key is valid ✓", "mode": "online"}
        except AuthenticationError:
            return {"valid": False, "key_type": "openai", "message": "Invalid OpenAI API key."}
        except RateLimitError:
            return {"valid": True, "key_type": "openai", "message": "OpenAI key valid but quota exceeded ⚠", "mode": "online"}
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"OpenAI connection error: {str(e)}")

    # ── Validate Gemini ──
    if key_type == "gemini":
        try:
            client = genai.Client(api_key=api_key)
            models = list(client.models.list())
            if not models:
                raise HTTPException(status_code=502, detail="Could not reach Gemini API.")
            return {"valid": True, "key_type": "gemini", "message": "Gemini API key is valid ✓", "mode": "gemini"}
        except Exception as e:
            err = str(e).lower()
            if "api_key" in err or "invalid" in err or "permission" in err or "403" in err:
                return {"valid": False, "key_type": "gemini", "message": "Invalid Gemini API key."}
            raise HTTPException(status_code=502, detail=f"Gemini connection error: {str(e)}")


# Keep old endpoints as aliases so existing frontend calls don't break
@app.post("/api/validate-openai-key")
async def validate_openai_key_alias(data: dict):
    return await validate_key(data)

@app.post("/api/validate-gemini-key")
async def validate_gemini_key_alias(data: dict):
    return await validate_key(data)





@app.get("/api/hardware")
def hardware_info():
    """Return hardware specs and suggested model."""
    hw = get_hardware_info()
    suggested = suggest_model(hw["ram_gb"])
    installed = get_installed_models()
    models_info = []
    for key, info in OLLAMA_MODELS.items():
        models_info.append({
            "name": key,
            "label": info["label"],
            "size": info["size"],
            "ram_required": info["ram_required"],
            "speed": info["speed"],
            "quality": info["quality"],
            "best_for": info["best_for"],
            "tier": info["tier"],
            "is_suggested": key == suggested,
            "is_installed": key in installed,
        })
    return {
        "hardware": hw,
        "suggested_model": suggested,
        "current_default": OLLAMA_MODEL_DEFAULT,
        "models": models_info,
        "installed_models": installed,
    }


@app.post("/api/pull-model")
async def pull_model(data: dict):
    """Pull/download a model from Ollama."""
    model = data.get("model", "")
    if model not in OLLAMA_MODELS:
        raise HTTPException(status_code=400, detail=f"Unknown model: {model}")
    try:
        r = requests.post(
            f"{OLLAMA_URL}/api/pull",
            json={"name": model, "stream": False},
            timeout=300,
        )
        r.raise_for_status()
        return {"success": True, "message": f"{model} downloaded successfully!"}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to pull model: {str(e)}")


@app.post("/api/regression/analyze")
async def regression_analyze(
    file: UploadFile = File(...),
    recent_modification_days: Optional[int] = Form(None),
    total_execution_days: int = Form(...),
    total_testers: int = Form(...),
    cases_per_tester_per_day: int = Form(...),
    mode: str = Form("offline"),
    openai_key: Optional[str] = Form(None),
    gemini_key: Optional[str] = Form(None),
    ollama_model: Optional[str] = Form(None),
):
    contents = await file.read()
    selected_model = ollama_model or OLLAMA_MODEL_DEFAULT
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    if len(df) > 2000:
        raise HTTPException(status_code=400, detail="Max 2000 test cases allowed.")
    if len(df) == 0:
        raise HTTPException(status_code=400, detail="Excel file is empty.")
    if mode == "online" and not openai_key:
        raise HTTPException(status_code=400, detail="OpenAI API key required for Online mode.")
    if mode == "gemini" and not gemini_key:
        raise HTTPException(status_code=400, detail="Gemini API key required for Gemini mode.")

    capacity = total_execution_days * total_testers * cases_per_tester_per_day

    rows = [row.to_dict() for _, row in df.iterrows()]
    loop = asyncio.get_event_loop()

    # Run all rows in parallel using ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=min(8, len(rows))) as pool:
        if mode == "online":
            futures = [loop.run_in_executor(pool, openai_analyze_risk, r, openai_key) for r in rows]
        elif mode == "gemini":
            futures = [loop.run_in_executor(pool, gemini_analyze_risk, r, gemini_key) for r in rows]
        else:
            futures = [loop.run_in_executor(pool, offline_analyze_risk, r, selected_model) for r in rows]
        results = await asyncio.gather(*futures)

    df["Risk Score"] = [r["risk_score"] for r in results]
    df["Priority"] = [r["priority"] for r in results]
    df["AI Risk Explanation"] = [r["explanation"] for r in results]
    df["AI Mode"] = [r.get("source", "ollama") for r in results]

    df_sorted = df.sort_values("Risk Score", ascending=False).reset_index(drop=True)
    df_sorted["Recommended for Execution"] = "No"
    if capacity > 0:
        df_sorted.loc[:capacity - 1, "Recommended for Execution"] = "Yes"

    session_id = str(uuid.uuid4())[:8]
    create_regression_excel(df_sorted, session_id, mode)

    p1 = len(df_sorted[df_sorted["Priority"] == "P1"])
    p2 = len(df_sorted[df_sorted["Priority"] == "P2"])
    p3 = len(df_sorted[df_sorted["Priority"] == "P3"])
    recommended_count = min(capacity, len(df_sorted))
    coverage_pct = round(recommended_count / len(df_sorted) * 100) if len(df_sorted) > 0 else 0

    return {
        "session_id": session_id,
        "mode": mode,
        "total_cases": len(df_sorted),
        "capacity": capacity,
        "p1_count": p1,
        "p2_count": p2,
        "p3_count": p3,
        "recommended_count": recommended_count,
        "coverage_percent": coverage_pct,
        "download_url": f"/api/download/{session_id}/regression",
        "ollama_model": selected_model
    }


@app.post("/api/automation/analyze")
async def automation_analyze(
    file: UploadFile = File(...),
    mode: str = Form("offline"),
    openai_key: Optional[str] = Form(None),
    gemini_key: Optional[str] = Form(None),
    ollama_model: Optional[str] = Form(None),
):
    contents = await file.read()
    selected_model = ollama_model or OLLAMA_MODEL_DEFAULT
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    if len(df) == 0:
        raise HTTPException(status_code=400, detail="Excel file is empty.")
    if mode == "online" and not openai_key:
        raise HTTPException(status_code=400, detail="OpenAI API key required for Online mode.")
    if mode == "gemini" and not gemini_key:
        raise HTTPException(status_code=400, detail="Gemini API key required for Gemini mode.")

    rows = [row.to_dict() for _, row in df.iterrows()]
    loop = asyncio.get_event_loop()

    # Run all rows in parallel using ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=min(8, len(rows))) as pool:
        if mode == "online":
            futures = [loop.run_in_executor(pool, openai_analyze_automation, r, openai_key) for r in rows]
        elif mode == "gemini":
            futures = [loop.run_in_executor(pool, gemini_analyze_automation, r, gemini_key) for r in rows]
        else:
            futures = [loop.run_in_executor(pool, offline_analyze_automation, r, selected_model) for r in rows]
        results = await asyncio.gather(*futures)

    df["Automation Suitability"] = [r["suitability"] for r in results]
    df["Confidence %"] = [r["confidence"] for r in results]
    df["AI Explanation"] = [r["explanation"] for r in results]
    df["AI Mode"] = [r.get("source", "ollama") for r in results]

    session_id = str(uuid.uuid4())[:8]
    create_automation_excel(df, session_id, mode)

    total = len(df)
    automatable = len(df[df["Automation Suitability"] == "Automatable"])
    partial = len(df[df["Automation Suitability"] == "Partial"])
    not_suitable = len(df[df["Automation Suitability"] == "Not Suitable"])

    return {
        "session_id": session_id,
        "mode": mode,
        "total_cases": total,
        "automatable_count": automatable,
        "partial_count": partial,
        "not_suitable_count": not_suitable,
        "automatable_percent": round(automatable / total * 100) if total else 0,
        "download_url": f"/api/download/{session_id}/automation",
        "ollama_model": selected_model
    }


@app.get("/api/download/{session_id}/{report_type}")
def download_report(session_id: str, report_type: str):
    filename = f"{report_type}_results_{session_id}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Report not found.")
    return FileResponse(
        path=path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )